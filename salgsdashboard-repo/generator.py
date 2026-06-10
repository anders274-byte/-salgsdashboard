#!/usr/bin/env python3
"""
Genererer index.html fra kildefilene.

Bruk:
    python3 generator.py <Salg_konvertert_output.xlsm> <FCA_Rapport.xlsx> [--dato YYYY-MM-DD]

Krever: pip install openpyxl

Kilder:
  - Salg_konvertert_output.xlsm : radnivå salg (bestilt/levert per dag/OLFI/MO).
                                  Filen er eks. direkteleveranser — verifisert mot
                                  stand-up-rapportens YTD-tall (eksakt match 2025 og 2026).
  - FCA_Rapport.xlsx            : «Daglig stand-up»-arket har direktelev YTD
                                  (PO til kundens DC) som info-tall.
"""
import json
import re
import sys
import datetime
from collections import defaultdict
from pathlib import Path

import openpyxl

HER = Path(__file__).parent


def les_direktelev(fca_fil):
    """Henter «Direktelev YTD <år> (info)»-raden fra Daglig stand-up-arket."""
    wb = openpyxl.load_workbook(fca_fil, read_only=True, data_only=True)
    ws = wb["Daglig stand-up"]
    funn = None
    tom = None
    for rad in ws.iter_rows(values_only=True):
        for v in rad:
            if isinstance(v, str) and (m := re.search(r"t\.o\.m\.\s*([\d]{2}\.[\d]{2})", v)):
                tom = m.group(1)
        celler = [v for v in rad if v is not None]
        for i, v in enumerate(celler):
            if isinstance(v, str) and v.startswith("Direktelev YTD"):
                tall = [x for x in celler[i:] if isinstance(x, (int, float))]
                tekst = [x for x in celler[i + 1:] if isinstance(x, str) and len(x) > 5]
                aar = re.search(r"(\d{4})", v)
                funn = {
                    "verdi": tall[0] if tall else None,
                    "aar": int(aar.group(1)) if aar else None,
                    "kommentar": tekst[-1] if tekst else "",
                }
    wb.close()
    if funn and tom and funn["aar"]:
        funn["perDato"] = f"{tom}.{funn['aar']}"
    return funn


def aggreger(salgsfil, idag):
    cutoff_md = (idag.month, idag.day)
    wb = openpyxl.load_workbook(salgsfil, read_only=True, data_only=True)
    ws = wb["Salg_konvertert"]

    daily = defaultdict(lambda: [0.0, 0.0])
    weekly = defaultdict(lambda: defaultdict(float))
    monthly = defaultdict(lambda: defaultdict(float))
    yearly = defaultdict(lambda: defaultdict(float))
    ytd = defaultdict(lambda: defaultdict(float))
    ktype_ytd = defaultdict(lambda: [0.0, 0.0])
    kat_ytd = defaultdict(lambda: [0.0, 0.0])
    vare_ytd = defaultdict(lambda: [0.0, 0.0, ""])
    forecast_w = defaultdict(float)

    for r in ws.iter_rows(min_row=2, values_only=True):
        d = r[1]
        if d is None:
            continue
        b, l = float(r[12] or 0), float(r[13] or 0)
        iy, iw, _ = d.isocalendar()
        if r[6] == "Rullerende":  # fremtidig/rullerende ordreinngang, kun bestilt
            forecast_w[(iy, iw)] += b
            continue
        olfi = int(r[11]) if r[11] else 0
        mo = r[21] or "?"
        kat = r[20] if r[20] and r[20] != "#N/A" else "UKJENT"

        daily[d.date()][0] += b
        daily[d.date()][1] += l
        for agg, key in ((weekly, (iy, iw)), (monthly, (d.year, d.month)), (yearly, d.year)):
            a = agg[key]
            a["b"] += b
            a["l"] += l
            if mo == "NG":
                a["ngB"] += b
                a["ngL"] += l
            else:
                a["reB"] += b
                a["reL"] += l
        if (d.month, d.day) <= cutoff_md:
            a = ytd[d.year]
            a["b"] += b
            a["l"] += l
            if mo == "NG":
                a["ngB"] += b
                a["ngL"] += l
            else:
                a["reB"] += b
                a["reL"] += l
            if d.year == idag.year:
                ktype_ytd[r[19]][0] += b
                ktype_ytd[r[19]][1] += l
                kat_ytd[kat][0] += b
                kat_ytd[kat][1] += l
                v = vare_ytd[olfi]
                v[0] += b
                v[1] += l
                v[2] = r[16] or str(olfi)
    wb.close()

    rnd = lambda x: round(x, 1) if x != int(x) else int(x)
    pakk = lambda dd: {k: rnd(v) for k, v in dd.items()}
    return {
        "generert": idag.strftime("%d.%m.%Y"),
        "idag": idag.strftime("%Y-%m-%d"),
        "aar": idag.year,
        "daily": [{"d": k.isoformat(), "b": rnd(v[0]), "l": rnd(v[1])} for k, v in sorted(daily.items())],
        "weekly": [{"y": k[0], "w": k[1], **pakk(v)} for k, v in sorted(weekly.items())],
        "monthly": [{"y": k[0], "m": k[1], **pakk(v)} for k, v in sorted(monthly.items())],
        "yearly": [{"y": k, **pakk(v)} for k, v in sorted(yearly.items())],
        "ytd": [{"y": k, **pakk(v)} for k, v in sorted(ytd.items())],
        "forecastWeekly": [{"y": k[0], "w": k[1], "b": rnd(v)} for k, v in sorted(forecast_w.items())],
        "kundetypeYtd": [{"kt": k, "b": rnd(v[0]), "l": rnd(v[1])} for k, v in sorted(ktype_ytd.items())],
        "kategoriYtd": [{"kat": k, "b": rnd(v[0]), "l": rnd(v[1])}
                        for k, v in sorted(kat_ytd.items(), key=lambda x: -x[1][1])],
        "toppVarer": [{"olfi": k, "navn": v[2], "b": rnd(v[0]), "l": rnd(v[1])}
                      for k, v in sorted(vare_ytd.items(), key=lambda x: -x[1][1])[:15]],
    }


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if len(args) < 2:
        sys.exit(__doc__)
    idag = datetime.date.today()
    for a in sys.argv[1:]:
        if a.startswith("--dato"):
            idag = datetime.date.fromisoformat(sys.argv[sys.argv.index(a) + 1] if a == "--dato" else a.split("=", 1)[1])

    data = aggreger(args[0], idag)
    data["direktelev"] = les_direktelev(args[1])

    mal = (HER / "template.html").read_text(encoding="utf-8")
    chartjs = (HER / "vendor" / "chart.umd.js").read_text(encoding="utf-8")
    json_str = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    html = mal.replace("/*__CHARTJS__*/", chartjs).replace("__DATA__", json_str)
    ut = HER / "index.html"
    ut.write_text(html, encoding="utf-8")
    print(f"Skrev {ut} ({ut.stat().st_size // 1024} kB) — {len(data['daily'])} dager, "
          f"{len(data['weekly'])} uker, direktelev: {data['direktelev']}")


if __name__ == "__main__":
    main()
